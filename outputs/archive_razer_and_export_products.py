import json
import math
import os
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = ROOT / ".env"
OUTPUT_ROOT = ROOT / "outputs"
IST = ZoneInfo("Asia/Kolkata")
PAGE_SIZE = 1000
RAZER_PATTERN = re.compile(r"(^|[^a-z0-9])razer([^a-z0-9]|$)", re.IGNORECASE)


def load_env():
    values = {}
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def request_json(method, url, service_key, **kwargs):
    headers = kwargs.pop("headers", {})
    headers.update(
        {
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
        }
    )
    response = requests.request(method, url, headers=headers, timeout=60, **kwargs)
    if not response.ok:
        raise RuntimeError(f"{method} {url} failed: {response.status_code} {response.text[:500]}")
    if not response.text:
        return None, response
    return response.json(), response


def fetch_all_products(supabase_url, service_key):
    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/products"
    products = []
    offset = 0
    total = None

    while True:
        params = {"select": "*", "order": "id.asc"}
        headers = {
            "Range-Unit": "items",
            "Range": f"{offset}-{offset + PAGE_SIZE - 1}",
            "Prefer": "count=exact",
        }
        rows, response = request_json("GET", endpoint, service_key, params=params, headers=headers)
        rows = rows or []
        products.extend(rows)
        content_range = response.headers.get("content-range") or response.headers.get("Content-Range")
        if total is None and content_range and "/" in content_range:
            try:
                total = int(content_range.rsplit("/", 1)[1])
            except ValueError:
                total = None
        if len(rows) < PAGE_SIZE:
            break
        offset += PAGE_SIZE

    return products, total or len(products)


def stringify(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(stringify(item) for item in value if stringify(item))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value)


def safe_excel_text(value, limit=32000):
    text = stringify(value).replace("\x00", "")
    return text if len(text) <= limit else text[: limit - 20] + " [truncated]"


def derived_price(row):
    for key in ("selling_price", "price"):
        value = row.get(key)
        if value is None or value == "":
            continue
        try:
            return float(value)
        except (TypeError, ValueError):
            continue
    return 0.0


def row_search_text(row):
    fields = [
        row.get("title"),
        row.get("brand"),
        row.get("sku"),
        row.get("category"),
        row.get("source_url"),
        row.get("description"),
        row.get("search_keywords"),
    ]
    return " ".join(stringify(field) for field in fields if stringify(field))


def is_razer_product(row):
    return bool(RAZER_PATTERN.search(row_search_text(row)))


def chunked(items, size):
    for index in range(0, len(items), size):
        yield items[index : index + size]


def archive_products(supabase_url, service_key, ids):
    if not ids:
        return []
    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/products"
    updated = []
    for group in chunked(ids, 80):
        id_filter = ",".join(group)
        url = f"{endpoint}?id=in.({id_filter})&select=id,title,brand,status,updated_at"
        rows, _ = request_json(
            "PATCH",
            url,
            service_key,
            headers={"Prefer": "return=representation"},
            json={"status": "archived"},
        )
        updated.extend(rows or [])
    return updated


def normalized_status(row):
    return str(row.get("status") or "").strip().lower()


def sort_product_key(row):
    return (
        safe_excel_text(row.get("category")).lower(),
        safe_excel_text(row.get("brand")).lower(),
        safe_excel_text(row.get("title")).lower(),
        safe_excel_text(row.get("sku")).lower(),
    )


def set_title(ws, title, subtitle, last_column):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    ws.cell(1, 1, title)
    ws.cell(2, 1, subtitle)
    ws.cell(1, 1).font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    ws.cell(2, 1).font = Font(name="Aptos", size=10, color="E7EEF8")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")
    for row in (1, 2):
        for col in range(1, last_column + 1):
            ws.cell(row, col).fill = PatternFill("solid", fgColor="1F3A5F")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20


def style_header_row(ws, row, last_column, fill="D9EAF7", color="17324D"):
    for col in range(1, last_column + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Aptos", size=10, bold=True, color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="9DB5CC"))


def write_matrix(ws, start_row, start_col, rows):
    for row_offset, row_values in enumerate(rows):
        for col_offset, value in enumerate(row_values):
            ws.cell(start_row + row_offset, start_col + col_offset, value)


def apply_sheet_basics(ws):
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.font = cell.font.copy(name="Aptos", size=10)
            cell.alignment = Alignment(vertical="top")


def add_table(ws, name, start_row, start_col, end_row, end_col):
    if end_row <= start_row:
        return
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build_workbook(products, archived_now, already_archived_razer, excluded_archived_count, output_path):
    generated_at = datetime.now(IST)
    export_rows = sorted(
        [row for row in products if normalized_status(row) != "archived"],
        key=sort_product_key,
    )
    razer_audit_rows = sorted(archived_now + already_archived_razer, key=sort_product_key)
    category_counts = Counter(safe_excel_text(row.get("category")) or "Uncategorized" for row in export_rows)
    brand_counts = Counter(safe_excel_text(row.get("brand")) or "Unknown" for row in export_rows)
    status_counts = Counter(safe_excel_text(row.get("status")) or "Unknown" for row in export_rows)
    total_value = sum(derived_price(row) * float(row.get("stock") or 0) for row in export_rows)

    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    set_title(
        summary,
        "LapKart Product Export",
        "Non-archived products after archiving Razer matches",
        6,
    )
    summary_rows = [
        ["Generated at", generated_at.strftime("%Y-%m-%d %H:%M:%S %Z")],
        ["Products exported", len(export_rows)],
        ["Archived products excluded", excluded_archived_count],
        ["Razer products archived now", len(archived_now)],
        ["Razer products already archived", len(already_archived_razer)],
        ["Razer rows in Products sheet", 0],
        ["Estimated stock value", total_value],
        ["Sort order", "Category, Brand, Title, SKU"],
    ]
    write_matrix(summary, 4, 1, summary_rows)
    summary["A4"].font = Font(name="Aptos", bold=True)
    for row in range(4, 4 + len(summary_rows)):
        summary.cell(row, 1).fill = PatternFill("solid", fgColor="EEF4FA")
        summary.cell(row, 1).font = Font(name="Aptos", bold=True, color="17324D")
        summary.cell(row, 2).border = Border(bottom=Side(style="thin", color="D9E4EE"))
    summary["B10"].number_format = '₹#,##0.00'
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 32

    summary.cell(4, 4, "Status")
    summary.cell(4, 5, "Count")
    style_header_row(summary, 4, 4 + 1, fill="D9EAF7")
    status_data = [[status, count] for status, count in sorted(status_counts.items())]
    write_matrix(summary, 5, 4, status_data)
    summary.column_dimensions["D"].width = 18
    summary.column_dimensions["E"].width = 12

    product_headers = [
        "ID",
        "Title",
        "Brand",
        "Category",
        "SKU",
        "Status",
        "Display Price",
        "MRP",
        "Stock",
        "Warranty",
        "Compatibility",
        "Highlights",
        "Search Keywords",
        "Primary Image",
        "Image Count",
        "Source URL",
        "Cost Price",
        "Max Discount %",
        "COD Allowed",
        "Returnable",
        "Universal",
        "Authenticity",
        "Condition",
        "GST Rate",
        "DOA Days",
        "Local Delivery",
        "Weight kg",
        "Dimensions cm",
        "Created At",
        "Updated At",
    ]
    product_values = []
    for row in export_rows:
        dimensions = " x ".join(
            safe_excel_text(row.get(key))
            for key in ("length_cm", "breadth_cm", "height_cm")
            if safe_excel_text(row.get(key))
        )
        product_values.append(
            [
                safe_excel_text(row.get("id")),
                safe_excel_text(row.get("title")),
                safe_excel_text(row.get("brand")),
                safe_excel_text(row.get("category")),
                safe_excel_text(row.get("sku")),
                safe_excel_text(row.get("status")),
                derived_price(row),
                row.get("mrp") or 0,
                row.get("stock") or 0,
                safe_excel_text(row.get("warranty")),
                safe_excel_text(row.get("compatibility")),
                safe_excel_text(row.get("highlights")),
                safe_excel_text(row.get("search_keywords")),
                safe_excel_text(row.get("image")),
                len(row.get("images") or []),
                safe_excel_text(row.get("source_url")),
                row.get("cost_price") or 0,
                row.get("max_discount_pct") or 0,
                stringify(row.get("cod_allowed")),
                stringify(row.get("returnable")),
                stringify(row.get("is_universal")),
                safe_excel_text(row.get("authenticity_grade")),
                safe_excel_text(row.get("condition_grade")),
                row.get("gst_rate") or 0,
                row.get("doa_policy_days") or 0,
                stringify(row.get("local_delivery_eligible")),
                row.get("weight_kg") or "",
                dimensions,
                safe_excel_text(row.get("created_at")),
                safe_excel_text(row.get("updated_at")),
            ]
        )

    products_ws = wb.create_sheet("Products")
    set_title(
        products_ws,
        "Products",
        "All non-archived LapKart products. Razer matches are archived before export.",
        len(product_headers),
    )
    write_matrix(products_ws, 4, 1, [product_headers] + product_values)
    style_header_row(products_ws, 4, len(product_headers))
    products_ws.freeze_panes = "A5"
    add_table(products_ws, "ProductsTable", 4, 1, 4 + len(product_values), len(product_headers))
    products_ws.conditional_formatting.add(
        f"I5:I{max(5, 4 + len(product_values))}",
        CellIsRule(operator="lessThanOrEqual", formula=["5"], fill=PatternFill("solid", fgColor="FFF2CC")),
    )
    products_ws.conditional_formatting.add(
        f"I5:I{max(5, 4 + len(product_values))}",
        CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor="F4CCCC")),
    )

    widths = {
        "A": 34,
        "B": 46,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 12,
        "G": 14,
        "H": 12,
        "I": 10,
        "J": 20,
        "K": 46,
        "L": 34,
        "M": 34,
        "N": 42,
        "O": 12,
        "P": 42,
        "Q": 12,
        "R": 14,
        "S": 12,
        "T": 12,
        "U": 12,
        "V": 16,
        "W": 14,
        "X": 10,
        "Y": 10,
        "Z": 14,
        "AA": 12,
        "AB": 18,
        "AC": 22,
        "AD": 22,
    }
    for col, width in widths.items():
        products_ws.column_dimensions[col].width = width
    for row in products_ws.iter_rows(min_row=5, max_row=products_ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (2, 11, 12, 13))
    for col in ("G", "H", "Q"):
        for cell in products_ws[col][4:]:
            cell.number_format = '₹#,##0.00'
    for col in ("I", "O", "R", "X", "Y"):
        for cell in products_ws[col][4:]:
            cell.number_format = "0"

    def summary_sheet(name, headers, rows, table_name):
        ws = wb.create_sheet(name)
        set_title(ws, name, "Counts from the Products sheet", len(headers))
        write_matrix(ws, 4, 1, [headers] + rows)
        style_header_row(ws, 4, len(headers))
        ws.freeze_panes = "A5"
        add_table(ws, table_name, 4, 1, 4 + len(rows), len(headers))
        for col_index, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_index)].width = 28 if col_index == 1 else 14
        return ws

    category_rows = [[category, count] for category, count in category_counts.most_common()]
    brand_rows = [[brand, count] for brand, count in brand_counts.most_common()]
    status_rows = [[status, count] for status, count in status_counts.most_common()]
    summary_sheet("Category Summary", ["Category", "Products"], category_rows, "CategorySummaryTable")
    summary_sheet("Brand Summary", ["Brand", "Products"], brand_rows, "BrandSummaryTable")
    summary_sheet("Status Summary", ["Status", "Products"], status_rows, "StatusSummaryTable")

    audit_headers = ["ID", "Title", "Brand", "Previous Status", "Archived In This Run", "SKU", "Category"]
    audit_values = [
        [
            safe_excel_text(row.get("id")),
            safe_excel_text(row.get("title")),
            safe_excel_text(row.get("brand")),
            safe_excel_text(row.get("status")),
            "Yes" if normalized_status(row) != "archived" else "No",
            safe_excel_text(row.get("sku")),
            safe_excel_text(row.get("category")),
        ]
        for row in razer_audit_rows
    ]
    audit_ws = wb.create_sheet("Removed Razer Audit")
    set_title(audit_ws, "Removed Razer Audit", "Products matched as Razer and excluded from Products", len(audit_headers))
    write_matrix(audit_ws, 4, 1, [audit_headers] + audit_values)
    style_header_row(audit_ws, 4, len(audit_headers), fill="FCE4D6", color="7A2E0E")
    audit_ws.freeze_panes = "A5"
    add_table(audit_ws, "RemovedRazerAuditTable", 4, 1, 4 + len(audit_values), len(audit_headers))
    for col, width in {"A": 34, "B": 48, "C": 18, "D": 16, "E": 18, "F": 18, "G": 18}.items():
        audit_ws.column_dimensions[col].width = width

    for ws in wb.worksheets:
        apply_sheet_basics(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "generated_at": generated_at.isoformat(),
        "output_path": str(output_path),
        "exported_products": len(export_rows),
        "archived_excluded": excluded_archived_count,
        "razer_archived_now": len(archived_now),
        "razer_already_archived": len(already_archived_razer),
        "category_count": len(category_counts),
        "brand_count": len(brand_counts),
        "status_counts": dict(status_counts),
    }


def verify_workbook(path, expected_export_count):
    wb = load_workbook(path, read_only=True, data_only=False)
    required_sheets = {
        "Summary",
        "Products",
        "Category Summary",
        "Brand Summary",
        "Status Summary",
        "Removed Razer Audit",
    }
    missing = required_sheets.difference(wb.sheetnames)
    if missing:
        raise RuntimeError(f"Workbook missing sheets: {sorted(missing)}")
    ws = wb["Products"]
    headers = [cell.value for cell in ws[4]]
    status_col = headers.index("Status") + 1
    title_col = headers.index("Title") + 1
    brand_col = headers.index("Brand") + 1
    keywords_col = headers.index("Search Keywords") + 1
    source_col = headers.index("Source URL") + 1
    data_rows = max(0, ws.max_row - 4)
    if data_rows != expected_export_count:
        raise RuntimeError(f"Workbook row count mismatch: {data_rows} != {expected_export_count}")
    archived_rows = []
    razer_rows = []
    for row_index in range(5, ws.max_row + 1):
        status = str(ws.cell(row_index, status_col).value or "").lower()
        if status == "archived":
            archived_rows.append(row_index)
        text = " ".join(
            str(ws.cell(row_index, col).value or "")
            for col in (title_col, brand_col, keywords_col, source_col)
        )
        if RAZER_PATTERN.search(text):
            razer_rows.append(row_index)
    if archived_rows:
        raise RuntimeError(f"Products sheet contains archived rows: {archived_rows[:5]}")
    if razer_rows:
        raise RuntimeError(f"Products sheet contains Razer rows: {razer_rows[:5]}")
    return {
        "sheets": wb.sheetnames,
        "products_rows": data_rows,
        "archived_rows_in_products": len(archived_rows),
        "razer_rows_in_products": len(razer_rows),
    }


def main():
    env = load_env()
    supabase_url = env.get("SUPABASE_URL") or env.get("PUBLIC_SUPABASE_URL")
    service_key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required in .env")

    before_products, before_total = fetch_all_products(supabase_url, service_key)
    razer_matches = [row for row in before_products if is_razer_product(row)]
    active_razer = [row for row in razer_matches if normalized_status(row) != "archived"]
    already_archived_razer = [row for row in razer_matches if normalized_status(row) == "archived"]
    updated = archive_products(supabase_url, service_key, [row["id"] for row in active_razer])

    after_products, after_total = fetch_all_products(supabase_url, service_key)
    non_archived = [row for row in after_products if normalized_status(row) != "archived"]
    remaining_razer = [row for row in non_archived if is_razer_product(row)]
    if remaining_razer:
        sample = [{"id": row.get("id"), "title": row.get("title")} for row in remaining_razer[:5]]
        raise RuntimeError(f"Razer products remain non-archived: {sample}")

    timestamp = datetime.now(IST).strftime("%Y%m%d-%H%M%S")
    output_dir = OUTPUT_ROOT / f"product-export-{timestamp}"
    output_path = output_dir / "lapkart-products-non-archived-after-razer-removal.xlsx"
    metadata = build_workbook(
        after_products,
        archived_now=active_razer,
        already_archived_razer=already_archived_razer,
        excluded_archived_count=after_total - len(non_archived),
        output_path=output_path,
    )
    verification = verify_workbook(output_path, len(non_archived))
    manifest = {
        "before_total": before_total,
        "after_total": after_total,
        "razer_matches_total": len(razer_matches),
        "razer_archived_now": len(updated),
        "razer_already_archived": len(already_archived_razer),
        "remaining_non_archived_razer": len(remaining_razer),
        "non_archived_exported": len(non_archived),
        "workbook": metadata,
        "verification": verification,
    }
    (output_dir / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
