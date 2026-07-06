import json
import os
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = ROOT / ".env"
OUTPUT_ROOT = ROOT / "outputs"
IST = ZoneInfo("Asia/Kolkata")
PAGE_SIZE = 1000

FRIENDLY_COLUMNS = [
    "id",
    "title",
    "brand",
    "category",
    "sku",
    "status",
    "price",
    "selling_price",
    "mrp",
    "cost_price",
    "stock",
    "warranty",
    "compatibility",
    "highlights",
    "search_keywords",
    "description",
    "image",
    "images",
    "source_url",
    "authenticity_grade",
    "condition_grade",
    "gst_rate",
    "hsn_code",
    "max_discount_pct",
    "cod_allowed",
    "returnable",
    "is_universal",
    "local_delivery_eligible",
    "doa_policy_days",
    "weight_kg",
    "length_cm",
    "breadth_cm",
    "height_cm",
    "rating",
    "review_count",
    "clearance",
    "specifications",
    "created_at",
    "updated_at",
]

TEXT_COLUMNS = {
    "id",
    "title",
    "brand",
    "category",
    "sku",
    "warranty",
    "compatibility",
    "highlights",
    "search_keywords",
    "description",
    "image",
    "images",
    "source_url",
    "authenticity_grade",
    "condition_grade",
    "hsn_code",
    "specifications",
    "created_at",
    "updated_at",
}

PRICE_COLUMNS = {"price", "selling_price", "mrp", "cost_price"}
INTEGER_COLUMNS = {"stock", "doa_policy_days", "review_count"}
DECIMAL_COLUMNS = {"gst_rate", "max_discount_pct", "weight_kg", "length_cm", "breadth_cm", "height_cm", "rating"}
BOOLEAN_COLUMNS = {"cod_allowed", "returnable", "is_universal", "local_delivery_eligible", "clearance"}


def load_env():
    values = dict(os.environ)
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith("#") or "=" not in stripped:
                continue
            key, value = stripped.split("=", 1)
            values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def request_json(method, url, service_key, **kwargs):
    headers = kwargs.pop("headers", {})
    headers.update(
        {
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
        }
    )
    response = requests.request(method, url, headers=headers, timeout=90, **kwargs)
    if not response.ok:
        raise RuntimeError(f"{method} {url} failed: {response.status_code} {response.text[:500]}")
    return response.json() if response.text else None, response


def fetch_all_products(supabase_url, service_key):
    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/products"
    products = []
    offset = 0
    total = None

    while True:
        headers = {
            "Range-Unit": "items",
            "Range": f"{offset}-{offset + PAGE_SIZE - 1}",
            "Prefer": "count=exact",
        }
        rows, response = request_json(
            "GET",
            endpoint,
            service_key,
            params={"select": "*", "order": "category.asc,brand.asc,title.asc,id.asc"},
            headers=headers,
        )
        rows = rows or []
        products.extend(rows)

        content_range = response.headers.get("content-range") or response.headers.get("Content-Range")
        if total is None and content_range and "/" in content_range:
            try:
                total = int(content_range.rsplit("/", 1)[1])
            except ValueError:
                total = None

        if len(rows) < PAGE_SIZE:
            break
        offset += PAGE_SIZE

    return products, total or len(products)


def stringify(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, list):
        parts = []
        for item in value:
            text = stringify(item)
            if text:
                parts.append(text)
        return "\n".join(parts)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def safe_cell(value, limit=32000):
    value = stringify(value)
    if isinstance(value, str):
        value = value.replace("\x00", "")
        return value if len(value) <= limit else value[: limit - 20] + " [truncated]"
    return value


def normalized_number(value):
    if value in ("", None):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def normalized_int(value):
    if value in ("", None):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return value


def cell_value(row, column):
    raw = row.get(column)
    if column in PRICE_COLUMNS or column in DECIMAL_COLUMNS:
        return normalized_number(raw)
    if column in INTEGER_COLUMNS:
        return normalized_int(raw)
    return safe_cell(raw)


def collect_columns(products):
    seen = set()
    for row in products:
        seen.update(row.keys())
    ordered = [column for column in FRIENDLY_COLUMNS if column in seen]
    ordered.extend(sorted(seen.difference(ordered)))
    return ordered


def display_header(column):
    return column.replace("_", " ").title()


def add_title(ws, title, subtitle, last_column):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    ws.cell(1, 1, title)
    ws.cell(2, 1, subtitle)
    for row_index in (1, 2):
        for col_index in range(1, last_column + 1):
            cell = ws.cell(row_index, col_index)
            cell.fill = PatternFill("solid", fgColor="1F3A5F")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(1, 1).font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    ws.cell(2, 1).font = Font(name="Aptos", size=10, color="E7EEF8")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22


def style_header(ws, row_index, last_column, fill="D9EAF7", color="17324D"):
    for col_index in range(1, last_column + 1):
        cell = ws.cell(row_index, col_index)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Aptos", size=10, bold=True, color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="9DB5CC"))
    ws.row_dimensions[row_index].height = 34


def add_table(ws, name, start_row, start_col, end_row, end_col):
    if end_row <= start_row:
        return
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_sheet_basics(ws):
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.font = cell.font.copy(name="Aptos", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=cell.alignment.wrap_text)


def set_column_widths(ws, columns):
    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        if column == "title":
            width = 48
        elif column in {"description", "compatibility", "highlights", "search_keywords", "specifications"}:
            width = 42
        elif column in {"image", "images", "source_url"}:
            width = 44
        elif column in {"id", "created_at", "updated_at"}:
            width = 34
        elif column in PRICE_COLUMNS or column in INTEGER_COLUMNS or column in DECIMAL_COLUMNS or column in BOOLEAN_COLUMNS:
            width = 14
        else:
            width = 20
        ws.column_dimensions[letter].width = width


def add_validations(ws, columns, product_count, lookups):
    if product_count <= 0:
        return
    first_row = 5
    last_row = first_row + product_count - 1
    column_map = {column: index + 1 for index, column in enumerate(columns)}

    def range_for(column):
        letter = get_column_letter(column_map[column])
        return f"{letter}{first_row}:{letter}{last_row}"

    validation_specs = {
        "status": "Lookups!$A$2:$A$20",
        "category": "Lookups!$C$2:$C$300",
        "brand": "Lookups!$E$2:$E$500",
        "authenticity_grade": "Lookups!$G$2:$G$50",
        "condition_grade": "Lookups!$I$2:$I$50",
    }
    for column, formula in validation_specs.items():
        if column in column_map and lookups.get(column):
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(range_for(column))

    yes_no = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yes_no)
    for column in BOOLEAN_COLUMNS.intersection(column_map):
        yes_no.add(range_for(column))


def build_workbook(products, total, output_path):
    generated_at = datetime.now(IST)
    columns = collect_columns(products)
    rows = [[cell_value(product, column) for column in columns] for product in products]

    status_counts = Counter(str(row.get("status") or "blank") for row in products)
    category_counts = Counter(str(row.get("category") or "Uncategorized") for row in products)
    brand_counts = Counter(str(row.get("brand") or "Unknown") for row in products)
    active_count = sum(1 for row in products if str(row.get("status") or "").lower() == "active")
    archived_count = sum(1 for row in products if str(row.get("status") or "").lower() == "archived")
    low_stock_count = sum(1 for row in products if isinstance(normalized_int(row.get("stock")), int) and normalized_int(row.get("stock")) <= 5)
    total_stock_value = 0.0
    for row in products:
        price = normalized_number(row.get("selling_price") if row.get("selling_price") is not None else row.get("price"))
        stock = normalized_int(row.get("stock"))
        if isinstance(price, (int, float)) and isinstance(stock, int):
            total_stock_value += price * stock

    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Start Here")
    add_title(summary, "LapKart Product Export", "Editable full product export from Supabase products table", 6)
    summary_rows = [
        ["Generated at", generated_at.strftime("%Y-%m-%d %H:%M:%S %Z")],
        ["Supabase rows fetched", total],
        ["Workbook product rows", len(products)],
        ["Active products", active_count],
        ["Archived products", archived_count],
        ["Low or zero stock rows", low_stock_count],
        ["Estimated stock value", total_stock_value],
        ["Main edit sheet", "Products Editable"],
        ["Important", "This export is read-only against the database. Editing this workbook does not update Supabase by itself."],
    ]
    for r, values in enumerate(summary_rows, start=4):
        for c, value in enumerate(values, start=1):
            summary.cell(r, c, value)
    for r in range(4, 4 + len(summary_rows)):
        summary.cell(r, 1).fill = PatternFill("solid", fgColor="EEF4FA")
        summary.cell(r, 1).font = Font(name="Aptos", bold=True, color="17324D")
        summary.cell(r, 2).border = Border(bottom=Side(style="thin", color="D9E4EE"))
    summary["B10"].number_format = '₹#,##0.00'
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 80

    summary.cell(4, 4, "Status")
    summary.cell(4, 5, "Count")
    style_header(summary, 4, 5)
    for r, (status, count) in enumerate(sorted(status_counts.items()), start=5):
        summary.cell(r, 4, status)
        summary.cell(r, 5, count)
    summary.column_dimensions["D"].width = 20
    summary.column_dimensions["E"].width = 12

    products_ws = wb.create_sheet("Products Editable")
    add_title(products_ws, "Products Editable", "All product rows and columns. Use filters, edit cells, then save as needed.", len(columns))
    products_ws.cell(4, 1)
    for col_index, column in enumerate(columns, start=1):
        products_ws.cell(4, col_index, display_header(column))
    for row_index, row_values in enumerate(rows, start=5):
        for col_index, value in enumerate(row_values, start=1):
            products_ws.cell(row_index, col_index, value)
    style_header(products_ws, 4, len(columns))
    products_ws.freeze_panes = "A5"
    add_table(products_ws, "ProductsEditableTable", 4, 1, 4 + len(rows), len(columns))
    set_column_widths(products_ws, columns)

    column_map = {column: index + 1 for index, column in enumerate(columns)}
    for column in PRICE_COLUMNS.intersection(column_map):
        for cell in products_ws[get_column_letter(column_map[column])][4:]:
            cell.number_format = '₹#,##0.00'
    for column in INTEGER_COLUMNS.intersection(column_map):
        for cell in products_ws[get_column_letter(column_map[column])][4:]:
            cell.number_format = "0"
    for column in DECIMAL_COLUMNS.intersection(column_map):
        for cell in products_ws[get_column_letter(column_map[column])][4:]:
            cell.number_format = "0.00"
    for column in TEXT_COLUMNS.intersection(column_map):
        for cell in products_ws[get_column_letter(column_map[column])][4:]:
            cell.alignment = Alignment(vertical="top", wrap_text=column in {"description", "compatibility", "highlights", "search_keywords", "images", "specifications"})

    if "stock" in column_map and rows:
        stock_letter = get_column_letter(column_map["stock"])
        stock_range = f"{stock_letter}5:{stock_letter}{4 + len(rows)}"
        products_ws.conditional_formatting.add(
            stock_range,
            CellIsRule(operator="lessThanOrEqual", formula=["5"], fill=PatternFill("solid", fgColor="FFF2CC")),
        )
        products_ws.conditional_formatting.add(
            stock_range,
            CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor="F4CCCC")),
        )

    lookups = {
        "status": sorted({str(row.get("status") or "") for row in products if row.get("status") not in (None, "")}),
        "category": sorted({str(row.get("category") or "") for row in products if row.get("category") not in (None, "")}),
        "brand": sorted({str(row.get("brand") or "") for row in products if row.get("brand") not in (None, "")}),
        "authenticity_grade": sorted({str(row.get("authenticity_grade") or "") for row in products if row.get("authenticity_grade") not in (None, "")}),
        "condition_grade": sorted({str(row.get("condition_grade") or "") for row in products if row.get("condition_grade") not in (None, "")}),
    }
    lookups_ws = wb.create_sheet("Lookups")
    lookup_columns = [
        ("status", "Status Values"),
        ("category", "Categories"),
        ("brand", "Brands"),
        ("authenticity_grade", "Authenticity"),
        ("condition_grade", "Condition"),
    ]
    add_title(lookups_ws, "Lookups", "Dropdown values used by Products Editable", 10)
    for offset, (key, label) in enumerate(lookup_columns):
        col = offset * 2 + 1
        lookups_ws.cell(4, col, label)
        for r, value in enumerate(lookups[key], start=5):
            lookups_ws.cell(r, col, value)
        lookups_ws.column_dimensions[get_column_letter(col)].width = 28
    style_header(lookups_ws, 4, 10)
    add_validations(products_ws, columns, len(rows), lookups)

    def add_summary_sheet(sheet_name, first_header, counter, table_name):
        ws = wb.create_sheet(sheet_name)
        add_title(ws, sheet_name, "Counts from Products Editable", 2)
        ws.cell(4, 1, first_header)
        ws.cell(4, 2, "Products")
        for row_index, (name, count) in enumerate(counter.most_common(), start=5):
            ws.cell(row_index, 1, name)
            ws.cell(row_index, 2, count)
        style_header(ws, 4, 2)
        ws.freeze_panes = "A5"
        add_table(ws, table_name, 4, 1, 4 + len(counter), 2)
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 14

    add_summary_sheet("Category Summary", "Category", category_counts, "CategorySummaryTable")
    add_summary_sheet("Brand Summary", "Brand", brand_counts, "BrandSummaryTable")
    add_summary_sheet("Status Summary", "Status", status_counts, "StatusSummaryTable")

    raw_ws = wb.create_sheet("Raw JSON")
    add_title(raw_ws, "Raw JSON", "One exact JSON object per product row for auditing/re-import mapping", 3)
    raw_ws.cell(4, 1, "ID")
    raw_ws.cell(4, 2, "Title")
    raw_ws.cell(4, 3, "Raw Product JSON")
    for row_index, product in enumerate(products, start=5):
        raw_ws.cell(row_index, 1, safe_cell(product.get("id")))
        raw_ws.cell(row_index, 2, safe_cell(product.get("title")))
        raw_ws.cell(row_index, 3, safe_cell(json.dumps(product, ensure_ascii=False, sort_keys=True)))
    style_header(raw_ws, 4, 3)
    raw_ws.freeze_panes = "A5"
    add_table(raw_ws, "RawJsonTable", 4, 1, 4 + len(products), 3)
    raw_ws.column_dimensions["A"].width = 34
    raw_ws.column_dimensions["B"].width = 48
    raw_ws.column_dimensions["C"].width = 120

    for ws in wb.worksheets:
        set_sheet_basics(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "generated_at": generated_at.isoformat(),
        "output_path": str(output_path),
        "products": len(products),
        "supabase_total": total,
        "columns": columns,
        "status_counts": dict(status_counts),
        "category_count": len(category_counts),
        "brand_count": len(brand_counts),
        "low_stock_count": low_stock_count,
    }


def verify_workbook(path, expected_rows, expected_columns):
    wb = load_workbook(path, read_only=False, data_only=False)
    required = {"Start Here", "Products Editable", "Lookups", "Category Summary", "Brand Summary", "Status Summary", "Raw JSON"}
    missing = required.difference(wb.sheetnames)
    if missing:
        raise RuntimeError(f"Missing sheets: {sorted(missing)}")
    ws = wb["Products Editable"]
    row_count = max(0, ws.max_row - 4)
    if row_count != expected_rows:
        raise RuntimeError(f"Products row count mismatch: {row_count} != {expected_rows}")
    headers = [cell.value for cell in ws[4]]
    if len(headers) != len(expected_columns):
        raise RuntimeError(f"Column count mismatch: {len(headers)} != {len(expected_columns)}")
    if "Id" not in headers or "Title" not in headers:
        raise RuntimeError("Expected ID and Title columns missing")
    return {
        "sheets": wb.sheetnames,
        "products_rows": row_count,
        "products_columns": len(headers),
        "first_product_title": ws["B5"].value if row_count else None,
        "has_table": len(ws.tables) > 0,
        "freeze_panes": ws.freeze_panes,
    }


def main():
    env = load_env()
    supabase_url = env.get("SUPABASE_URL") or env.get("PUBLIC_SUPABASE_URL") or env.get("VITE_SUPABASE_URL")
    service_key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise RuntimeError("SUPABASE_URL/PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required")

    products, total = fetch_all_products(supabase_url, service_key)
    timestamp = datetime.now(IST).strftime("%Y%m%d-%H%M%S")
    output_dir = OUTPUT_ROOT / f"product-export-all-editable-{timestamp}"
    output_path = output_dir / "lapkart-all-products-editable.xlsx"
    metadata = build_workbook(products, total, output_path)
    verification = verify_workbook(output_path, len(products), metadata["columns"])
    manifest = {
        "metadata": metadata,
        "verification": verification,
    }
    (output_dir / "verification.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
