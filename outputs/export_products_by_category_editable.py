import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = ROOT / ".env"
OUTPUT_ROOT = ROOT / "outputs"
IST = ZoneInfo("Asia/Kolkata")
PAGE_SIZE = 1000

EDIT_COLUMNS = [
    "id",
    "title",
    "part_numbers",
    "brand",
    "category",
    "sku",
    "status",
    "stock",
    "selling_price",
    "price",
    "mrp",
    "cost_price",
    "warranty",
    "compatibility",
    "highlights",
    "search_keywords",
    "description",
    "image",
    "images",
    "source_url",
    "authenticity_grade",
    "condition_grade",
    "gst_rate",
    "hsn_code",
    "max_discount_pct",
    "cod_allowed",
    "returnable",
    "is_universal",
    "local_delivery_eligible",
    "doa_policy_days",
    "weight_kg",
    "length_cm",
    "breadth_cm",
    "height_cm",
    "rating",
    "clearance",
    "specifications",
    "created_at",
    "updated_at",
]

EDITABLE_COLUMNS = {
    "title",
    "part_numbers",
    "brand",
    "category",
    "sku",
    "status",
    "stock",
    "selling_price",
    "price",
    "mrp",
    "cost_price",
    "warranty",
    "compatibility",
    "highlights",
    "search_keywords",
    "description",
    "image",
    "images",
    "source_url",
    "authenticity_grade",
    "condition_grade",
    "gst_rate",
    "hsn_code",
    "max_discount_pct",
    "cod_allowed",
    "returnable",
    "is_universal",
    "local_delivery_eligible",
    "doa_policy_days",
    "weight_kg",
    "length_cm",
    "breadth_cm",
    "height_cm",
    "clearance",
    "specifications",
}

SYSTEM_COLUMNS = {"id", "created_at", "updated_at", "rating"}
PRICE_COLUMNS = {"selling_price", "price", "mrp", "cost_price"}
INTEGER_COLUMNS = {"stock", "doa_policy_days"}
DECIMAL_COLUMNS = {"gst_rate", "max_discount_pct", "weight_kg", "length_cm", "breadth_cm", "height_cm", "rating"}
BOOLEAN_COLUMNS = {"cod_allowed", "returnable", "is_universal", "local_delivery_eligible", "clearance"}
WRAP_COLUMNS = {"title", "part_numbers", "compatibility", "highlights", "search_keywords", "description", "images", "specifications"}

CATEGORY_LABELS = {
    "displays": "Displays",
    "ics": "ICs",
    "batteries": "Batteries",
    "chargers": "Chargers",
    "keyboards": "Keyboards",
    "speakers": "Speakers",
    "dc_jacks": "DC Jacks",
    "flex_cables": "Flex Cables",
    "cooling": "Cooling",
    "power_buttons": "Power Buttons",
    "processors": "Processors",
    "ssd": "SSD",
    "palmrests": "Palmrests",
}


def load_env():
    values = dict(os.environ)
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
            stripped = line.strip()
            if stripped and not stripped.startswith("#") and "=" in stripped:
                key, value = stripped.split("=", 1)
                values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def request_json(method, url, service_key, **kwargs):
    headers = kwargs.pop("headers", {})
    headers.update(
        {
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
        }
    )
    response = requests.request(method, url, headers=headers, timeout=90, **kwargs)
    if not response.ok:
        raise RuntimeError(f"{method} {url} failed: {response.status_code} {response.text[:500]}")
    return response.json() if response.text else None, response


def fetch_all_products(supabase_url, service_key):
    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/products"
    products = []
    offset = 0
    total = None
    while True:
        rows, response = request_json(
            "GET",
            endpoint,
            service_key,
            params={"select": "*", "order": "category.asc,brand.asc,title.asc,id.asc"},
            headers={
                "Range-Unit": "items",
                "Range": f"{offset}-{offset + PAGE_SIZE - 1}",
                "Prefer": "count=exact",
            },
        )
        rows = rows or []
        products.extend(rows)
        content_range = response.headers.get("content-range") or response.headers.get("Content-Range")
        if total is None and content_range and "/" in content_range:
            try:
                total = int(content_range.rsplit("/", 1)[1])
            except ValueError:
                total = None
        if len(rows) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
    return products, total or len(products)


def stringify(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, list):
        return "\n".join(filter(None, (str(stringify(item)) for item in value)))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def safe_cell(value, limit=32000):
    value = stringify(value)
    if isinstance(value, str):
        value = value.replace("\x00", "")
        return value if len(value) <= limit else value[: limit - 20] + " [truncated]"
    return value


def add_part_number(parts, value):
    if value in (None, ""):
        return
    values = value if isinstance(value, list) else [value]
    for item in values:
        text = str(item or "").strip()
        if not text:
            continue
        for candidate in re.split(r"[,;/\n]+", text):
            candidate = candidate.strip(" .;:-")
            if is_part_number(candidate) and candidate not in parts:
                parts.append(candidate)


def add_first_part_number(parts, value):
    if value in (None, ""):
        return
    for candidate in re.split(r"[,;/\n]+", str(value)):
        candidate = candidate.strip(" .;:-")
        lower = candidate.lower()
        if any(noise in lower for noise in ("sparepartworld", "anti-glare", "non-touch", "touchscreen")):
            continue
        if is_part_number(candidate) and candidate not in parts:
            parts.append(candidate)
            return


def is_part_number(candidate):
    if not candidate:
        return False
    normalized = candidate.strip()
    lower = normalized.lower()
    if len(normalized) < 5 or len(normalized) > 40:
        return False
    if lower in {"display", "screen", "laptop", "compatible", "panel", "edp", "lvds", "hd", "fhd", "qhd", "uhd"}:
        return False
    if re.fullmatch(r"\d{2}(\.\d)?", normalized):
        return False
    if re.fullmatch(r"\d{3,4}x\d{3,4}", lower):
        return False
    if re.fullmatch(r"\d+\s*-?\s*pin", lower):
        return False
    has_digit = any(char.isdigit() for char in normalized)
    has_letter = any(char.isalpha() for char in normalized)
    has_part_separator = any(char in normalized for char in ".-_/")
    compact = re.sub(r"\s+", "", normalized)
    if re.fullmatch(r"\d{6,}", compact):
        return True
    if re.fullmatch(r"[A-Z0-9]{5,}", compact) and (has_digit or normalized == normalized.upper()):
        return True
    return has_digit and (has_letter or has_part_separator)


def extract_display_part_numbers(row):
    if str(row.get("category") or "").lower() != "displays":
        return ""

    parts = []
    specs = row.get("specifications")
    if isinstance(specs, dict):
        for key in ("Panel Part Number", "Primary Part Number", "Sub-Partnumbers"):
            add_part_number(parts, specs.get(key))
        table = specs.get("Panel Part Number Table")
        if isinstance(table, list):
            for entry in table:
                if isinstance(entry, dict):
                    add_part_number(parts, entry.get("value"))

    compatibility = str(row.get("compatibility") or "")
    match = re.search(r"Panel part numbers?:\s*(.*?)(?:\.\s*(?:Compatible|Match)|$)", compatibility, flags=re.IGNORECASE)
    if match:
        add_part_number(parts, match.group(1))
    elif not parts:
        generic_match = re.search(r"\bPart numbers?:\s*(.*?)(?:\.\s*(?:Compatible|Match)|$)", compatibility, flags=re.IGNORECASE)
        if generic_match:
            add_first_part_number(parts, generic_match.group(1))

    title = str(row.get("title") or "")
    title_match = re.search(
        r"\bDisplay\s+(.+?)(?:\s+\d{2}(?:\.\d)?\s*(?:in|inch)|\s+\d{3,4}x\d{3,4}|$)",
        title,
        flags=re.IGNORECASE,
    )
    if title_match:
        add_part_number(parts, title_match.group(1))

    return "\n".join(parts)


def number_or_blank(value, integer=False):
    if value in ("", None):
        return None
    try:
        numeric = float(value)
        return int(numeric) if integer else numeric
    except (TypeError, ValueError):
        return value


def value_for(row, column):
    if column == "part_numbers":
        return extract_display_part_numbers(row)
    if column in PRICE_COLUMNS or column in DECIMAL_COLUMNS:
        return number_or_blank(row.get(column))
    if column in INTEGER_COLUMNS:
        return number_or_blank(row.get(column), integer=True)
    return safe_cell(row.get(column))


def display_header(column):
    return column.replace("_", " ").title()


def sheet_safe_name(name, used):
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", name).strip() or "Blank"
    cleaned = re.sub(r"\s+", " ", cleaned)[:31]
    base = cleaned
    counter = 2
    while cleaned in used:
        suffix = f" {counter}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(cleaned)
    return cleaned


def table_safe_name(name):
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", name)
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"
    return cleaned[:240]


def add_title(ws, title, subtitle, last_column):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    ws.cell(1, 1, title)
    ws.cell(2, 1, subtitle)
    for row_index in (1, 2):
        for col_index in range(1, last_column + 1):
            cell = ws.cell(row_index, col_index)
            cell.fill = PatternFill("solid", fgColor="17324D")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(1, 1).font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    ws.cell(2, 1).font = Font(name="Aptos", size=10, color="E7EEF8")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22


def style_header(ws, row_index, columns):
    for col_index, column in enumerate(columns, start=1):
        cell = ws.cell(row_index, col_index)
        cell.fill = PatternFill("solid", fgColor="FFF2CC" if column in EDITABLE_COLUMNS else "E7ECF3")
        cell.font = Font(name="Aptos", size=10, bold=True, color="17324D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="9DB5CC"))
    ws.row_dimensions[row_index].height = 34


def add_table(ws, name, start_row, start_col, end_row, end_col):
    if end_row <= start_row:
        return
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_safe_name(name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_widths(ws, columns):
    for index, column in enumerate(columns, start=1):
        if column == "title":
            width = 48
        elif column == "part_numbers":
            width = 30
        elif column in {"description", "compatibility", "highlights", "search_keywords", "specifications"}:
            width = 40
        elif column in {"image", "images", "source_url"}:
            width = 42
        elif column in {"id", "created_at", "updated_at"}:
            width = 32
        elif column in PRICE_COLUMNS or column in INTEGER_COLUMNS or column in DECIMAL_COLUMNS or column in BOOLEAN_COLUMNS:
            width = 13
        else:
            width = 18
        ws.column_dimensions[get_column_letter(index)].width = width


def format_body(ws, columns, row_count):
    if row_count <= 0:
        return
    first_row = 5
    last_row = first_row + row_count - 1
    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        cells = ws[f"{letter}{first_row}:{letter}{last_row}"]
        for row in cells:
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=column in WRAP_COLUMNS)
                cell.protection = Protection(locked=column in SYSTEM_COLUMNS)
                if column in SYSTEM_COLUMNS:
                    cell.fill = PatternFill("solid", fgColor="F4F7FA")
        if column in PRICE_COLUMNS:
            for row in cells:
                for cell in row:
                    cell.number_format = '₹#,##0.00'
        elif column in INTEGER_COLUMNS:
            for row in cells:
                for cell in row:
                    cell.number_format = "0"
        elif column in DECIMAL_COLUMNS:
            for row in cells:
                for cell in row:
                    cell.number_format = "0.00"
    if "stock" in columns:
        stock_col = get_column_letter(columns.index("stock") + 1)
        stock_range = f"{stock_col}{first_row}:{stock_col}{last_row}"
        ws.conditional_formatting.add(
            stock_range,
            CellIsRule(operator="lessThanOrEqual", formula=["5"], fill=PatternFill("solid", fgColor="FFF2CC")),
        )
        ws.conditional_formatting.add(
            stock_range,
            CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor="F4CCCC")),
        )


def apply_validations(ws, columns, row_count, lookup_refs):
    if row_count <= 0:
        return
    first_row = 5
    last_row = first_row + row_count - 1
    column_indexes = {column: index + 1 for index, column in enumerate(columns)}

    def add_list(column, formula):
        if column not in column_indexes:
            return
        letter = get_column_letter(column_indexes[column])
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letter}{first_row}:{letter}{last_row}")

    add_list("status", lookup_refs["status"])
    add_list("category", lookup_refs["category"])
    add_list("brand", lookup_refs["brand"])
    add_list("authenticity_grade", lookup_refs["authenticity_grade"])
    add_list("condition_grade", lookup_refs["condition_grade"])

    yes_no = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yes_no)
    for column in sorted(BOOLEAN_COLUMNS.intersection(column_indexes)):
        letter = get_column_letter(column_indexes[column])
        yes_no.add(f"{letter}{first_row}:{letter}{last_row}")


def build_product_sheet(wb, sheet_name, title, subtitle, products, columns, lookup_refs, table_name):
    ws = wb.create_sheet(sheet_name)
    add_title(ws, title, subtitle, len(columns))
    for col_index, column in enumerate(columns, start=1):
        ws.cell(4, col_index, display_header(column))
    for row_index, product in enumerate(products, start=5):
        for col_index, column in enumerate(columns, start=1):
            ws.cell(row_index, col_index, value_for(product, column))
    style_header(ws, 4, columns)
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    add_table(ws, table_name, 4, 1, 4 + len(products), len(columns))
    set_widths(ws, columns)
    format_body(ws, columns, len(products))
    apply_validations(ws, columns, len(products), lookup_refs)
    ws.protection.sheet = False
    return ws


def build_workbook(products, total, output_path):
    generated_at = datetime.now(IST)
    available_columns = [
        column for column in EDIT_COLUMNS if column == "part_numbers" or any(column in product for product in products)
    ]
    extras = sorted(set().union(*(product.keys() for product in products)).difference(available_columns))
    raw_columns = available_columns + extras
    category_groups = defaultdict(list)
    for product in products:
        category_groups[str(product.get("category") or "blank")].append(product)

    lookups = {
        "status": sorted({str(row.get("status") or "") for row in products if row.get("status") not in (None, "")}),
        "category": sorted(category_groups),
        "brand": sorted({str(row.get("brand") or "") for row in products if row.get("brand") not in (None, "")}),
        "authenticity_grade": sorted({str(row.get("authenticity_grade") or "") for row in products if row.get("authenticity_grade") not in (None, "")}),
        "condition_grade": sorted({str(row.get("condition_grade") or "") for row in products if row.get("condition_grade") not in (None, "")}),
    }

    wb = Workbook()
    wb.remove(wb.active)
    used_sheet_names = set()

    start = wb.create_sheet("Start Here")
    used_sheet_names.add("Start Here")
    add_title(start, "LapKart Editable Product Workbook", "Pick a category tab below, edit the yellow columns, keep ID unchanged for existing products.", 8)
    summary_rows = [
        ["Generated at", generated_at.strftime("%Y-%m-%d %H:%M:%S %Z")],
        ["Products exported", len(products)],
        ["Supabase rows fetched", total],
        ["How to edit", "Use category sheets like Displays, ICs, Batteries. Yellow headers are editable fields; grey ID/date/rating fields are reference fields."],
        ["Important", "This Excel file does not update Supabase automatically. It is prepared for easy editing/review/import mapping."],
    ]
    for r, values in enumerate(summary_rows, start=4):
        for c, value in enumerate(values, start=1):
            start.cell(r, c, value)
    for r in range(4, 4 + len(summary_rows)):
        start.cell(r, 1).fill = PatternFill("solid", fgColor="EEF4FA")
        start.cell(r, 1).font = Font(name="Aptos", bold=True, color="17324D")
        start.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    start.column_dimensions["A"].width = 24
    start.column_dimensions["B"].width = 100

    start.cell(11, 1, "Category")
    start.cell(11, 2, "Products")
    start.cell(11, 3, "Open Sheet")
    start.cell(11, 4, "Active")
    start.cell(11, 5, "Draft")
    for col in range(1, 6):
        start.cell(11, col).fill = PatternFill("solid", fgColor="D9EAF7")
        start.cell(11, col).font = Font(name="Aptos", bold=True, color="17324D")
        start.cell(11, col).alignment = Alignment(horizontal="center")

    lookup_refs = {}
    lookups_ws = wb.create_sheet("Lookups")
    used_sheet_names.add("Lookups")
    add_title(lookups_ws, "Lookups", "Dropdown values used throughout the workbook", 10)
    lookup_order = [
        ("status", "Status Values"),
        ("category", "Categories"),
        ("brand", "Brands"),
        ("authenticity_grade", "Authenticity"),
        ("condition_grade", "Condition"),
    ]
    for offset, (key, label) in enumerate(lookup_order):
        col = offset * 2 + 1
        lookups_ws.cell(4, col, label)
        for r, value in enumerate(lookups[key], start=5):
            lookups_ws.cell(r, col, value)
        letter = get_column_letter(col)
        last_row = max(5, 4 + len(lookups[key]))
        lookup_refs[key] = f"Lookups!${letter}$5:${letter}${last_row}"
        lookups_ws.column_dimensions[letter].width = 28
    style_header(lookups_ws, 4, ["status", "gap", "category", "gap", "brand", "gap", "authenticity_grade", "gap", "condition_grade", "gap"])
    lookups_ws.freeze_panes = "A5"
    lookups_ws.sheet_view.showGridLines = False

    all_sheet = build_product_sheet(
        wb,
        "All Products",
        "All Products",
        "Full editable product list. Use category sheets for easier focused editing.",
        products,
        available_columns,
        lookup_refs,
        "AllProductsTable",
    )
    used_sheet_names.add("All Products")

    category_rows = []
    sorted_categories = sorted(category_groups.items(), key=lambda item: (-len(item[1]), item[0]))
    for idx, (category, items) in enumerate(sorted_categories, start=12):
        label = CATEGORY_LABELS.get(category, category.replace("_", " ").title())
        sheet_name = sheet_safe_name(label, used_sheet_names)
        category_rows.append((category, label, sheet_name, len(items)))
        active = sum(1 for item in items if str(item.get("status") or "").lower() == "active")
        draft = sum(1 for item in items if str(item.get("status") or "").lower() == "draft")
        start.cell(idx, 1, label)
        start.cell(idx, 2, len(items))
        start.cell(idx, 3, f"Open {sheet_name}")
        start.cell(idx, 3).hyperlink = f"#'{sheet_name}'!A1"
        start.cell(idx, 3).style = "Hyperlink"
        start.cell(idx, 4, active)
        start.cell(idx, 5, draft)

        build_product_sheet(
            wb,
            sheet_name,
            label,
            f"{len(items)} products in category '{category}'. Edit yellow columns; keep ID unchanged.",
            items,
            available_columns,
            lookup_refs,
            f"{table_safe_name(sheet_name)}Table",
        )

    add_table(start, "CategoryIndexTable", 11, 1, 11 + len(category_rows), 5)
    start.freeze_panes = "A12"
    start.sheet_view.showGridLines = False
    for col, width in {"A": 26, "B": 12, "C": 28, "D": 12, "E": 12}.items():
        start.column_dimensions[col].width = width

    raw = build_product_sheet(
        wb,
        "Raw All Columns",
        "Raw All Columns",
        "All database columns. Use this only when you need technical fields not shown first elsewhere.",
        products,
        raw_columns,
        lookup_refs,
        "RawAllColumnsTable",
    )
    raw.sheet_properties.tabColor = "808080"
    all_sheet.sheet_properties.tabColor = "1F77B4"

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.font:
                    cell.font = cell.font.copy(name="Aptos")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "generated_at": generated_at.isoformat(),
        "output_path": str(output_path),
        "products": len(products),
        "supabase_total": total,
        "edit_columns": available_columns,
        "raw_columns": raw_columns,
        "categories": [
            {"category": category, "sheet": sheet_name, "products": count}
            for category, label, sheet_name, count in category_rows
        ],
        "status_counts": dict(Counter(str(row.get("status") or "blank") for row in products)),
    }


def verify(path, expected_rows, expected_categories):
    wb = load_workbook(path, read_only=False, data_only=False)
    required = {"Start Here", "Lookups", "All Products", "Raw All Columns"}
    missing = required.difference(wb.sheetnames)
    if missing:
        raise RuntimeError(f"Missing sheets: {sorted(missing)}")
    all_ws = wb["All Products"]
    all_rows = all_ws.max_row - 4
    if all_rows != expected_rows:
        raise RuntimeError(f"All Products row mismatch: {all_rows} != {expected_rows}")
    missing_category_sheets = [entry["sheet"] for entry in expected_categories if entry["sheet"] not in wb.sheetnames]
    if missing_category_sheets:
        raise RuntimeError(f"Missing category sheets: {missing_category_sheets}")
    sheet_counts = {}
    for entry in expected_categories:
        ws = wb[entry["sheet"]]
        rows = ws.max_row - 4
        if rows != entry["products"]:
            raise RuntimeError(f"{entry['sheet']} row mismatch: {rows} != {entry['products']}")
        sheet_counts[entry["sheet"]] = rows
    return {
        "sheet_count": len(wb.sheetnames),
        "sheets": wb.sheetnames,
        "all_products_rows": all_rows,
        "category_sheet_counts": sheet_counts,
        "all_products_freeze": all_ws.freeze_panes,
        "all_products_tables": list(all_ws.tables.keys()),
    }


def main():
    env = load_env()
    supabase_url = env.get("SUPABASE_URL") or env.get("PUBLIC_SUPABASE_URL") or env.get("VITE_SUPABASE_URL")
    service_key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise RuntimeError("SUPABASE_URL/PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required")
    products, total = fetch_all_products(supabase_url, service_key)
    timestamp = datetime.now(IST).strftime("%Y%m%d-%H%M%S")
    output_dir = OUTPUT_ROOT / f"product-export-by-category-editable-{timestamp}"
    output_path = output_dir / "lapkart-products-by-category-editable.xlsx"
    metadata = build_workbook(products, total, output_path)
    verification = verify(output_path, len(products), metadata["categories"])
    manifest = {"metadata": metadata, "verification": verification}
    (output_dir / "verification.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
